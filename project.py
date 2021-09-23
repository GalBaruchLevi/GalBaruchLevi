import math
# from re import S
# from typing import Counter
import pyodbc
# from os import read
import random
import openpyxl

 
class Customer:
    def __init__(self, ID, name ,budget,city):
        self.ID = ID
        self.name = name
        self.budget =int(budget)
        self.city = city
        self.SpendingScore= 0
        self.orders=[]
      
    def BuyProduct (self,Store,selection):
        x= True
        message = "Sorry, you dont have enough money to buy this product "
        if((self.SpendingScore >= Store.shop[selection-1].price)):
            useInSpenScore = input ("Do you want to use your scores? pleas enter y/n")
            if (useInSpenScore == "y"):
                self.SpendingScore -= Store.shop[selection-1].price
                self.orders.append(Store.shop[selection-1])
                
            elif (useInSpenScore == "f"):
                if (self.budget >=  Store.shop[selection-1].price):
                   enoughMoney(self,Store,selection)
                   self.orders.append(Store.shop[selection-1])
                else: 
                    x= False 
                    print(message)
        else:
            if(int(self.budget) >= Store.shop[selection-1].price):
                enoughMoney(self,Store,selection)
                self.orders.append(Store.shop[selection-1])
                
            else:
                x= False
                print(message)
        if(x):       
            write_csv('customerBuyData.txt', ' The product: ' + Store.shop[selection-1].productName + ' With the product id: ' + str(Store.shop[selection-1].productCode) + ' has been bougth by '+ self.name)


    def lottery(self,Store,selection):
        randnum=random.randint(1,10)
        if (randnum == selection):
            self.SpendingScore += selection*100
            print('Congratulations, you won '+ str((Store.shop[selection-1]*10)+'$'))
        else:
            print("Sorry, you did not win the lottery, maybe another time ")    
    

class Product:
    def __init__ (self, productCode, productName, price):
        self.productCode = productCode
        self.productName =productName
        self.price = price
 
class Store:
    def __init__(self, shop):
        self.shop = shop

    def printInfo (self):
        num = 1
        print ("list of customers")
        for i in self.shop:
            print(f'{num} - code: {i.productCode} , name: {i.productName} , price: {i.price}')
            num+=1 

class Manager (Customer):
    def __init__(self, ID, name,budget, city):
        super().__init__(ID,name,budget,city)
        

def write_csv(path, table):
        f= open(path, 'a')
        f.write(table+'\n')
        f.close()


def enoughMoney(self,Store,selection):
    self.budget -= Store.shop[selection-1].price
    self.SpendingScore += (Store.shop[selection-1].price)*0.2
    print("enjoy your product:)")       


def printManagerMenu():
    print('1.Buy product\n2.View store purchases\n3.See the VIP members with budget above 100$\n4.Exit')

def ProductInfoXl (orders):
    wb = openpyxl.load_workbook(filename='gal.xlsx')
    sheet = wb['sheet1'] 

    for row in sheet['A1' : sheet.cell(row=sheet.max_row, column=sheet.max_column).coordinate]:
        for cell in row:
            cell.value = None
            
            start_row_index=2
            start_column_index=1

            rows=int(input("How much product you bogth?"))
            cols=3
            y=0
            
            first_cell_coordinates= sheet.cell(row=start_row_index, column=start_column_index).coordinate
            last_cell_coordinates=sheet.cell(row=start_row_index+int(rows)-1, column=start_column_index+cols-1).coordinate

            for i in sheet[first_cell_coordinates:last_cell_coordinates]:
                for cell in i:
                    
                            z=1
                            sheet.cell(row=start_row_index, column=z).value =orders[y].productCode
                            z+=1
                            sheet.cell(row=start_row_index, column=z).value =orders[y].productName
                            z+=1
                            sheet.cell(row=start_row_index, column=z).value =orders[y].price
                            z+=1
                start_row_index+=1 
                y+=1

                                
            sheet.cell(1,1).value="Product Code"
            sheet.cell(1,2).value="Product Name" 
            sheet.cell(1,3).value="Price"                         
                            
               
            wb.save('Transaction.xlsx')  

def ExcNoneQuery(sql_command, values):
    server= 'DESKTOP-GI7REBJ\SQLEXPRESS'
    database = 'Store1'
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server}; \
                          SERVER=' + server +'; \
                          DATABASE='+ database +'; \
                          Trusted_connection=yes;')
    crsr = cnxn.cursor()
    crsr.execute(sql_command, values)

    crsr.commit()
    crsr.close()
    cnxn.close()

def ExcQuery(sql_command,values):
    server= 'DESKTOP-GI7REBJ\SQLEXPRESS'
    database = 'Store1'
    cnxn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server}; \
                          SERVER=' + server +'; \
                          DATABASE='+ database +'; \
                          Trusted_connection=yes;')

    crsr = cnxn.cursor()
    crsr.execute(sql_command,values)
   

    table = []
    for row in crsr:
        table.append (row)
       
   
    crsr.close()
    cnxn.close()
    return table

def printMenu():
        print("1.Print info \n2.Buy product\n3.See your Spending Score\n4.See your total purchase\n5.Save your transaction in excel file\n6.Press to join to VIP club\n7.Insert your transaction to db\n8.Exit")

products = [Product(111,"shirt",100),Product(222,"dress",200),Product(333,"pants",160),Product(444,"skirt",140),
Product(555,"hoodie",250),Product(666,"jacket",270),Product(777,"swimsuit",190),
Product(888,"hat",80),Product(999,"shoes",300)]

shopping = Store(products)

ifmanger = input("Are you a manegr? Press y or n  ")

if(ifmanger == "n"):
   
    ID = input ("Please enter your ID ")
    name = input("Please enter your name ")
    budget = input ("Please enter your budget ")
    city = input ("Please enter your city ")
    
    custo =Customer(ID, name, budget, city)


    flag = True
    while (flag):
        printMenu()
        select =int(input())

        if select == 1:
            shopping.printInfo()

        elif select == 2:
            print("Let's start shopping! \n choose the product that you want to buy")
            shopping.printInfo()  
            selection1 = int(input("please choose a product: ")) 
            custo.BuyProduct(shopping,selection1)
            custo.lottery(shopping,selection1)
            
        elif select == 3:  
            print(custo.SpendingScore)
        
        elif select==4:
            print ("Here is your recepit: ")
            counter=0
            for i in custo.orders:
                counter+=i.price
                print ('product name: '+ i.productName +'\nprice: '+ str(i.price) +'$ ')
               
            print('********\nyour total purchase: '+ str(counter) +'$' )

        elif select == 5:
            ProductInfoXl(custo.orders)
                   
        elif select == 6:
            sqlCommand= 'Insert into Customer (ID,name,budget,city,SpendingScoure) VALUES(?,?,?,?,?)'
            values = (custo.ID, custo.name, custo.budget,custo.city,custo.SpendingScore)
            ExcNoneQuery(sqlCommand, values)
            print("Congratulations, your joining the VIP Club has been completed successfully,\n your details have been saved in the database\n6.Exit")
        elif select == 7:
            if (len(custo.orders)>0):
                sqlCommand= 'Insert into Product (productCode,productName,price) VALUES(?,?,?)'
                for i in custo.orders: 
                    values = (i.productCode,i.productName,i.price)
                ExcNoneQuery(sqlCommand, values)
                print("Saved successfully")
            else:
                print("Sorry, you need to buy something first")    

        elif select == 8:
            print("bye,bye")
            flag=False
        else:
            print("invalid input")    

else:
    manager1 = Manager(1111,'Nir',100000,"Israel")
    flag2=True
    while(flag2): 
        
        printManagerMenu()
        Mselect =int(input())

        if(Mselect==1):
            selection2 =int(input("please choose a product: ")) 
            manager1.BuyProduct(shopping, selection2)


        elif(Mselect==2):
            f=open('customerBuyData.txt')
            content = f.read()
            print(content)
            f.close()

        elif(Mselect==3):
            sql_Command='Select name , budget from Customer Where budget > ? order by ID'
            values= (100)
            tab= ExcQuery(sql_Command,values)
            for i in tab:
                print (i)
          
        elif(Mselect==4):
            print("Bye,Bye...")
            flag2=False
        else:
            print("invalid input")      
           
